from __future__ import annotations

import csv
import io
from collections.abc import Callable
from dataclasses import dataclass
from datetime import datetime
from xml.sax.saxutils import escape as xml_escape

from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.core.pdf_utils import logo_flowable
from app.core.sanitize import escape_formula
from app.schemas.commerce_dashboard import CommerceDashboardOut


@dataclass
class ExportMeta:
    company_name: str
    logo_url: str | None
    report_title: str
    filter_label: str
    generated_at: datetime


@dataclass
class WidgetSpec:
    title: str
    headers: list[str]
    rows: Callable[[CommerceDashboardOut], list[list[str]]]
    numeric_from: int


def _rows_kpis(data: CommerceDashboardOut) -> list[list[str]]:
    k = data.kpis
    return [
        ["Total Online Orders", str(k.total_online_orders)],
        ["Swiggy Orders", str(k.swiggy_orders)],
        ["Zomato Orders", str(k.zomato_orders)],
        ["Total Online Revenue", f"{k.total_online_revenue:.2f}"],
        ["Average Order Value", f"{k.average_order_value:.2f}"],
        ["Cancelled Orders", str(k.cancelled_orders)],
        ["Unmapped SKUs", str(k.unmapped_sku_count)],
    ]


def _rows_channel_comparison(data: CommerceDashboardOut) -> list[list[str]]:
    return [
        [c.platform.capitalize(), f"{c.revenue:.2f}", str(c.orders), f"{c.average_order_value:.2f}", str(c.cancelled_orders)]
        for c in data.channel_comparison
    ]


def _rows_revenue_trend(data: CommerceDashboardOut) -> list[list[str]]:
    return [
        [p.bucket_label, f"{p.swiggy_revenue:.2f}", f"{p.zomato_revenue:.2f}", f"{p.total_revenue:.2f}", str(p.order_count)]
        for p in data.revenue_trend
    ]


def _rows_top_products(data: CommerceDashboardOut) -> list[list[str]]:
    return [[p.product_name, f"{p.quantity_sold:g}", f"{p.revenue:.2f}"] for p in data.top_products]


WIDGET_SPECS: dict[str, WidgetSpec] = {
    "kpis": WidgetSpec("KPIs", ["Metric", "Value"], _rows_kpis, 1),
    "channel_comparison": WidgetSpec(
        "Channel Comparison", ["Platform", "Revenue", "Orders", "Avg Order Value", "Cancelled"], _rows_channel_comparison, 1
    ),
    "revenue_trend": WidgetSpec(
        "Revenue Trend", ["Date", "Swiggy Revenue", "Zomato Revenue", "Total Revenue", "Orders"], _rows_revenue_trend, 1
    ),
    "top_products": WidgetSpec("Top Selling Products", ["Product", "Qty Sold", "Revenue"], _rows_top_products, 1),
}

_TABLE_STYLE = TableStyle(
    [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d1d5db")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f9fafb")]),
    ]
)


def _sheet_rows(spec: WidgetSpec, data: CommerceDashboardOut) -> list[list[str]]:
    rows = []
    for row in spec.rows(data):
        text_columns, rest = row[: spec.numeric_from], row[spec.numeric_from :]
        rows.append([escape_formula(v) for v in text_columns] + rest)
    return rows


def export_widget_excel(widget: str, data: CommerceDashboardOut) -> bytes:
    spec = WIDGET_SPECS[widget]
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = spec.title[:31]
    sheet.append(spec.headers)
    for row in _sheet_rows(spec, data):
        sheet.append(row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def export_widget_csv(widget: str, data: CommerceDashboardOut) -> bytes:
    spec = WIDGET_SPECS[widget]
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(spec.headers)
    for row in _sheet_rows(spec, data):
        writer.writerow(row)
    return output.getvalue().encode("utf-8")


def _pdf_header_story(meta: ExportMeta, styles) -> list:
    story: list = []
    if meta.logo_url:
        logo = logo_flowable(meta.logo_url, max_width=30 * mm, max_height=15 * mm)
        if logo:
            story.append(logo)
            story.append(Spacer(1, 2 * mm))
    story.append(Paragraph(xml_escape(meta.company_name), styles["Heading2"]))
    story.append(Paragraph(xml_escape(meta.report_title), styles["Title"]))
    story.append(Paragraph(f"Period: {xml_escape(meta.filter_label)}", styles["Normal"]))
    story.append(Paragraph(f"Generated: {meta.generated_at.strftime('%Y-%m-%d %H:%M')}", styles["Normal"]))
    story.append(Spacer(1, 4 * mm))
    return story


def _pdf_footer(canvas, doc) -> None:
    canvas.saveState()
    canvas.setFont("Helvetica", 8)
    canvas.setFillColor(colors.HexColor("#6b7280"))
    canvas.drawString(14 * mm, 10 * mm, f"Generated {datetime.now():%Y-%m-%d %H:%M}")
    canvas.drawRightString(landscape(A4)[0] - 14 * mm, 10 * mm, f"Page {doc.page}")
    canvas.restoreState()


def _widget_table_story(spec: WidgetSpec, data: CommerceDashboardOut, styles) -> list:
    story: list = [Paragraph(spec.title, styles["Heading2"]), Spacer(1, 2 * mm)]
    rows = spec.rows(data)
    if not rows:
        story.append(Paragraph("No data available for the selected period.", styles["Normal"]))
    else:
        table = Table([spec.headers] + rows, repeatRows=1)
        style = TableStyle(_TABLE_STYLE.getCommands())
        style.add("ALIGN", (spec.numeric_from, 0), (-1, -1), "RIGHT")
        table.setStyle(style)
        story.append(table)
    story.append(Spacer(1, 6 * mm))
    return story


def export_widget_pdf(widget: str, data: CommerceDashboardOut, meta: ExportMeta) -> bytes:
    spec = WIDGET_SPECS[widget]
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=14 * mm, bottomMargin=18 * mm, leftMargin=14 * mm, rightMargin=14 * mm)
    styles = getSampleStyleSheet()
    story = _pdf_header_story(meta, styles) + _widget_table_story(spec, data, styles)
    doc.build(story, onFirstPage=_pdf_footer, onLaterPages=_pdf_footer)
    return buffer.getvalue()


def export_full_dashboard_excel(data: CommerceDashboardOut) -> bytes:
    workbook = Workbook()
    first = True
    for spec in WIDGET_SPECS.values():
        sheet = workbook.active if first else workbook.create_sheet()
        first = False
        sheet.title = spec.title[:31]
        sheet.append(spec.headers)
        for row in _sheet_rows(spec, data):
            sheet.append(row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def export_full_dashboard_pdf(data: CommerceDashboardOut, meta: ExportMeta) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=14 * mm, bottomMargin=18 * mm, leftMargin=14 * mm, rightMargin=14 * mm)
    styles = getSampleStyleSheet()
    story = _pdf_header_story(meta, styles)
    for spec in WIDGET_SPECS.values():
        story.extend(_widget_table_story(spec, data, styles))
    doc.build(story, onFirstPage=_pdf_footer, onLaterPages=_pdf_footer)
    return buffer.getvalue()
