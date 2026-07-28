from __future__ import annotations

import csv
import io
from collections.abc import Callable
from dataclasses import dataclass
from datetime import datetime
from xml.sax.saxutils import escape as xml_escape

from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.core.pdf_utils import logo_flowable
from app.core.sanitize import escape_formula
from app.schemas.analytics import DashboardOut, TrendComparisonOut


@dataclass
class ExportMeta:
    company_name: str
    logo_url: str | None
    report_title: str
    filter_label: str
    generated_at: datetime


def _rows_kpis(data: DashboardOut) -> list[list[str]]:
    k = data.kpis
    fields = [
        ("Total Sales", f"{k.total_sales:.2f}"),
        ("Net Sales", f"{k.net_sales:.2f}"),
        ("Total Orders", str(k.total_orders)),
        ("Average Order Value", f"{k.average_order_value:.2f}"),
        ("New Customers", str(k.new_customers)),
        ("Credit Sales", f"{k.credit_sales:.2f}"),
        ("Cash Sales", f"{k.cash_sales:.2f}"),
        ("Card Payments", f"{k.card_payments:.2f}"),
        ("UPI Payments", f"{k.upi_payments:.2f}"),
        ("Returned Orders", str(k.returned_orders)),
        ("Returned Amount", f"{k.returned_amount:.2f}"),
        ("Discount Amount", f"{k.discount_amount:.2f}"),
        ("Tax Collected", f"{k.tax_collected:.2f}"),
        ("Total Customers (as of today)", str(k.total_customers)),
        ("Total Products (as of today)", str(k.total_products)),
        ("Outstanding Amount (as of today)", f"{k.outstanding_amount:.2f}"),
    ]
    return [[label, value] for label, value in fields]


def _rows_sales_trend(data: DashboardOut) -> list[list[str]]:
    return [[p.bucket_label, f"{p.revenue:.2f}", str(p.order_count)] for p in data.sales_trend]


def _rows_hourly_sales(data: DashboardOut) -> list[list[str]]:
    return [[f"{p.hour:02d}:00", f"{p.revenue:.2f}", str(p.order_count)] for p in data.hourly_sales]


def _rows_payment_methods(data: DashboardOut) -> list[list[str]]:
    return [[p.method.upper(), f"{p.amount:.2f}", str(p.count)] for p in data.payment_methods]


def _rows_top_products(data: DashboardOut) -> list[list[str]]:
    return [[p.name, p.identifier_value, f"{p.qty_sold:g}", f"{p.revenue:.2f}"] for p in data.top_products]


def _rows_top_categories(data: DashboardOut) -> list[list[str]]:
    return [[c.name, f"{c.qty_sold:g}", f"{c.revenue:.2f}"] for c in data.top_categories]


def _rows_sales_by_employee(data: DashboardOut) -> list[list[str]]:
    return [[e.name, f"{e.revenue:.2f}", str(e.order_count)] for e in data.sales_by_employee]


def _rows_tax_breakdown(data: DashboardOut) -> list[list[str]]:
    return [[f"{t.tax_rate_percent:g}%", f"{t.taxable_amount:.2f}", f"{t.tax_amount:.2f}"] for t in data.tax_breakdown]


def _rows_discount_analysis(data: DashboardOut) -> list[list[str]]:
    d = data.discount_analysis
    return [
        ["Total Discount", f"{d.total_discount:.2f}"],
        ["Discounted Invoices", str(d.discounted_invoice_count)],
        ["Total Invoices", str(d.total_invoice_count)],
        ["Average Discount %", f"{d.avg_discount_percent:.2f}%"],
    ]


def _rows_recent_invoices(data: DashboardOut) -> list[list[str]]:
    return [
        [
            i.invoice_number,
            i.created_at.strftime("%Y-%m-%d %H:%M"),
            i.customer_name or "Walk-in",
            i.status,
            f"{i.total_amount:.2f}",
            i.payment_method.upper(),
        ]
        for i in data.recent_invoices
    ]


def _rows_customer_retention(data: DashboardOut) -> list[list[str]]:
    r = data.customer_retention
    return [
        ["New Customers", str(r.new_customers)],
        ["Returning Customers", str(r.returning_customers)],
        ["Retention Rate", f"{r.retention_rate_percent:.1f}%"],
    ]


def _rows_cash_flow_summary(data: DashboardOut) -> list[list[str]]:
    c = data.cash_flow_summary
    rows = [["Total Cash In", "", f"{c.total_cash_in:.2f}"]]
    for m in c.by_method:
        rows.append([f"By Method: {m.method.upper()}", str(m.count), f"{m.amount:.2f}"])
    for p in c.daily:
        rows.append([f"By Day: {p.bucket_label}", "", f"{p.amount:.2f}"])
    return rows


@dataclass
class WidgetSpec:
    title: str
    headers: list[str]
    rows: Callable[[DashboardOut], list[list[str]]]
    numeric_from: int  # column index from which cells are right-aligned and not formula-escaped


WIDGET_SPECS: dict[str, WidgetSpec] = {
    "kpis": WidgetSpec("KPIs", ["Metric", "Value"], _rows_kpis, 1),
    "sales_trend": WidgetSpec("Sales Trend", ["Period", "Revenue", "Orders"], _rows_sales_trend, 1),
    "hourly_sales": WidgetSpec("Hourly Sales", ["Hour", "Revenue", "Orders"], _rows_hourly_sales, 1),
    "payment_methods": WidgetSpec("Payment Methods", ["Method", "Amount", "Count"], _rows_payment_methods, 1),
    "top_products": WidgetSpec("Top Selling Products", ["Product", "Identifier", "Qty Sold", "Revenue"], _rows_top_products, 2),
    "top_categories": WidgetSpec("Top Categories", ["Category", "Qty Sold", "Revenue"], _rows_top_categories, 1),
    "sales_by_employee": WidgetSpec("Sales by Employee", ["Employee", "Revenue", "Orders"], _rows_sales_by_employee, 1),
    "tax_breakdown": WidgetSpec("Tax Breakdown", ["Tax Rate", "Taxable Amount", "Tax Amount"], _rows_tax_breakdown, 1),
    "discount_analysis": WidgetSpec("Discount Analysis", ["Metric", "Value"], _rows_discount_analysis, 1),
    "recent_invoices": WidgetSpec(
        "Recent Invoices", ["Invoice #", "Date", "Customer", "Status", "Total", "Payment Method"], _rows_recent_invoices, 4
    ),
    "customer_retention": WidgetSpec("Customer Retention", ["Metric", "Value"], _rows_customer_retention, 1),
    "cash_flow_summary": WidgetSpec("Cash Flow Summary", ["Item", "Count", "Amount"], _rows_cash_flow_summary, 1),
}

_TABLE_STYLE = TableStyle(
    [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d1d5db")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f9fafb")]),
    ]
)


def _sheet_rows(spec: WidgetSpec, data: DashboardOut) -> list[list[str]]:
    rows = []
    for row in spec.rows(data):
        text_columns, rest = row[: spec.numeric_from], row[spec.numeric_from :]
        rows.append([escape_formula(v) for v in text_columns] + rest)
    return rows


def export_widget_excel(widget: str, data: DashboardOut) -> bytes:
    spec = WIDGET_SPECS[widget]
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = spec.title[:31]
    sheet.append(spec.headers)
    for row in _sheet_rows(spec, data):
        sheet.append(row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def export_widget_csv(widget: str, data: DashboardOut) -> bytes:
    spec = WIDGET_SPECS[widget]
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(spec.headers)
    for row in _sheet_rows(spec, data):
        writer.writerow(row)
    return output.getvalue().encode("utf-8")


def _pdf_header_story(meta: ExportMeta, styles) -> list:
    story: list = []
    if meta.logo_url:
        logo = logo_flowable(meta.logo_url, max_width=30 * mm, max_height=15 * mm)
        if logo:
            story.append(logo)
            story.append(Spacer(1, 2 * mm))
    story.append(Paragraph(xml_escape(meta.company_name), styles["Heading2"]))
    story.append(Paragraph(xml_escape(meta.report_title), styles["Title"]))
    story.append(Paragraph(f"Period: {xml_escape(meta.filter_label)}", styles["Normal"]))
    story.append(Paragraph(f"Generated: {meta.generated_at.strftime('%Y-%m-%d %H:%M')}", styles["Normal"]))
    story.append(Spacer(1, 4 * mm))
    return story


def _pdf_footer(canvas, doc) -> None:
    canvas.saveState()
    canvas.setFont("Helvetica", 8)
    canvas.setFillColor(colors.HexColor("#6b7280"))
    canvas.drawString(14 * mm, 10 * mm, f"Generated {datetime.now():%Y-%m-%d %H:%M}")
    canvas.drawRightString(landscape(A4)[0] - 14 * mm, 10 * mm, f"Page {doc.page}")
    canvas.restoreState()


def _widget_table_story(spec: WidgetSpec, data: DashboardOut, styles) -> list:
    story: list = [Paragraph(spec.title, styles["Heading2"]), Spacer(1, 2 * mm)]
    rows = spec.rows(data)
    if not rows:
        story.append(Paragraph("No data available for the selected period.", styles["Normal"]))
    else:
        table = Table([spec.headers] + rows, repeatRows=1)
        style = TableStyle(_TABLE_STYLE.getCommands())
        style.add("ALIGN", (spec.numeric_from, 0), (-1, -1), "RIGHT")
        table.setStyle(style)
        story.append(table)
    story.append(Spacer(1, 6 * mm))
    return story


def export_widget_pdf(widget: str, data: DashboardOut, meta: ExportMeta) -> bytes:
    spec = WIDGET_SPECS[widget]
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4), topMargin=14 * mm, bottomMargin=18 * mm, leftMargin=14 * mm, rightMargin=14 * mm
    )
    styles = getSampleStyleSheet()
    story = _pdf_header_story(meta, styles) + _widget_table_story(spec, data, styles)
    doc.build(story, onFirstPage=_pdf_footer, onLaterPages=_pdf_footer)
    return buffer.getvalue()


def export_full_dashboard_excel(data: DashboardOut) -> bytes:
    workbook = Workbook()
    first = True
    for spec in WIDGET_SPECS.values():
        sheet = workbook.active if first else workbook.create_sheet()
        first = False
        sheet.title = spec.title[:31]
        sheet.append(spec.headers)
        for row in _sheet_rows(spec, data):
            sheet.append(row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def export_full_dashboard_pdf(data: DashboardOut, meta: ExportMeta) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4), topMargin=14 * mm, bottomMargin=18 * mm, leftMargin=14 * mm, rightMargin=14 * mm
    )
    styles = getSampleStyleSheet()
    story = _pdf_header_story(meta, styles)
    for spec in WIDGET_SPECS.values():
        story.extend(_widget_table_story(spec, data, styles))
    doc.build(story, onFirstPage=_pdf_footer, onLaterPages=_pdf_footer)
    return buffer.getvalue()


_COMPARISON_HEADERS = ["Metric", "Current", "Previous", "Difference", "Growth %"]


def _rows_trend_comparison(result: TrendComparisonOut) -> list[list[str]]:
    return [
        [
            m.label,
            f"{m.current_value:.2f}",
            f"{m.previous_value:.2f}",
            f"{m.absolute_difference:+.2f}",
            f"{m.growth_percent:+.1f}%" if m.growth_percent is not None else "—",
        ]
        for m in result.metrics
    ]


def export_trend_comparison_excel(result: TrendComparisonOut) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Trend Comparison"
    sheet.append(_COMPARISON_HEADERS)
    for row in _rows_trend_comparison(result):
        text_columns, rest = row[:1], row[1:]
        sheet.append([escape_formula(v) for v in text_columns] + rest)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def export_trend_comparison_csv(result: TrendComparisonOut) -> bytes:
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(_COMPARISON_HEADERS)
    for row in _rows_trend_comparison(result):
        text_columns, rest = row[:1], row[1:]
        writer.writerow([escape_formula(v) for v in text_columns] + rest)
    return output.getvalue().encode("utf-8")


def export_trend_comparison_pdf(result: TrendComparisonOut, meta: ExportMeta) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4), topMargin=14 * mm, bottomMargin=18 * mm, leftMargin=14 * mm, rightMargin=14 * mm
    )
    styles = getSampleStyleSheet()
    story = _pdf_header_story(meta, styles)
    story.append(Paragraph(f"{result.current_label} vs {result.previous_label}", styles["Heading2"]))
    story.append(Spacer(1, 2 * mm))
    rows = _rows_trend_comparison(result)
    table = Table([_COMPARISON_HEADERS] + rows, repeatRows=1)
    style = TableStyle(_TABLE_STYLE.getCommands())
    style.add("ALIGN", (1, 0), (-1, -1), "RIGHT")
    table.setStyle(style)
    story.append(table)
    doc.build(story, onFirstPage=_pdf_footer, onLaterPages=_pdf_footer)
    return buffer.getvalue()
