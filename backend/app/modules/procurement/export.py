from __future__ import annotations

import csv
import io
from collections.abc import Callable
from dataclasses import dataclass
from datetime import datetime
from typing import Any
from xml.sax.saxutils import escape as xml_escape

from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.core.pdf_utils import logo_flowable
from app.core.sanitize import escape_formula
from app.schemas.procurement_analytics import ProcurementAnalyticsOut, ProcurementReportsOut


@dataclass
class ExportMeta:
    company_name: str
    logo_url: str | None
    report_title: str
    filter_label: str
    generated_at: datetime


@dataclass
class WidgetSpec:
    title: str
    headers: list[str]
    rows: Callable[[Any], list[list[str]]]
    numeric_from: int


# --- Analytics dashboard widgets ---


def _rows_analytics_kpis(data: ProcurementAnalyticsOut) -> list[list[str]]:
    k = data.kpis
    return [
        ["Total Purchase Value", f"{k.total_purchase_value:.2f}"],
        ["Total Purchases", str(k.total_purchase_count)],
        ["Total Returns Value", f"{k.total_return_value:.2f}"],
        ["Inventory Value", f"{k.inventory_value:.2f}"],
        ["Gross Profit (approx.)", f"{k.gross_profit:.2f}"],
        ["Gross Margin %", f"{k.gross_margin_percent:.2f}%"],
    ]


def _rows_purchase_trend(data: ProcurementAnalyticsOut) -> list[list[str]]:
    return [[p.bucket_label, f"{p.purchase_value:.2f}", str(p.purchase_count)] for p in data.purchase_trend]


def _rows_vendor_analysis(data: ProcurementAnalyticsOut) -> list[list[str]]:
    return [
        [v.vendor_name, f"{v.purchase_value:.2f}", str(v.purchase_count), f"{v.purchase_quantity:g}", f"{v.outstanding_amount:.2f}"]
        for v in data.vendor_analysis
    ]


def _rows_product_analysis(data: ProcurementAnalyticsOut) -> list[list[str]]:
    return [[p.product_name, f"{p.quantity_purchased:g}", f"{p.purchase_value:.2f}", f"{p.avg_cost_price:.2f}"] for p in data.product_analysis]


ANALYTICS_WIDGET_SPECS: dict[str, WidgetSpec] = {
    "kpis": WidgetSpec("KPIs", ["Metric", "Value"], _rows_analytics_kpis, 1),
    "purchase_trend": WidgetSpec("Purchase Trend", ["Date", "Purchase Value", "Purchase Count"], _rows_purchase_trend, 1),
    "vendor_analysis": WidgetSpec(
        "Vendor Analysis", ["Vendor", "Purchase Value", "Purchase Count", "Quantity", "Outstanding"], _rows_vendor_analysis, 1
    ),
    "product_analysis": WidgetSpec(
        "Product Analysis", ["Product", "Quantity Purchased", "Purchase Value", "Avg Cost Price"], _rows_product_analysis, 1
    ),
}

# --- Reports ---


def _rows_purchase_report(data: ProcurementReportsOut) -> list[list[str]]:
    return [
        [r.purchase_number, r.vendor_name, r.purchase_date.isoformat(), r.status, f"{r.total_amount:.2f}", r.payment_status, f"{r.outstanding_amount:.2f}"]
        for r in data.purchase_report
    ]


def _rows_vendor_report(data: ProcurementReportsOut) -> list[list[str]]:
    return [
        [v.vendor_name, "Active" if v.is_active else "Inactive", str(v.purchase_count), f"{v.total_purchase_value:.2f}", f"{v.total_outstanding:.2f}"]
        for v in data.vendor_report
    ]


def _rows_outstanding_vendor_report(data: ProcurementReportsOut) -> list[list[str]]:
    return [[v.vendor_name, v.phone or "", f"{v.outstanding_amount:.2f}", f"{v.credit_limit:.2f}"] for v in data.outstanding_vendor_report]


def _rows_purchase_return_report(data: ProcurementReportsOut) -> list[list[str]]:
    return [
        [r.return_number, r.purchase_number, r.vendor_name, r.status, f"{r.refund_amount:.2f}", r.created_at.strftime("%Y-%m-%d %H:%M")]
        for r in data.purchase_return_report
    ]


def _rows_gst_report(data: ProcurementReportsOut) -> list[list[str]]:
    return [[f"{g.tax_rate_percent:g}%", f"{g.taxable_amount:.2f}", f"{g.tax_amount:.2f}"] for g in data.gst_report]


REPORT_WIDGET_SPECS: dict[str, WidgetSpec] = {
    "purchase_report": WidgetSpec(
        "Purchase Report", ["Purchase #", "Vendor", "Date", "Status", "Total", "Payment Status", "Outstanding"], _rows_purchase_report, 4
    ),
    "vendor_report": WidgetSpec(
        "Vendor Report", ["Vendor", "Status", "Purchase Count", "Total Purchase Value", "Total Outstanding"], _rows_vendor_report, 2
    ),
    "outstanding_vendor_report": WidgetSpec(
        "Outstanding Vendor Report", ["Vendor", "Phone", "Outstanding", "Credit Limit"], _rows_outstanding_vendor_report, 2
    ),
    "purchase_return_report": WidgetSpec(
        "Purchase Return Report", ["Return #", "Purchase #", "Vendor", "Status", "Refund", "Date"], _rows_purchase_return_report, 4
    ),
    "gst_report": WidgetSpec("GST Report", ["Tax Rate", "Taxable Amount", "Tax Amount"], _rows_gst_report, 1),
}

ALL_WIDGET_SPECS: dict[str, WidgetSpec] = {**ANALYTICS_WIDGET_SPECS, **REPORT_WIDGET_SPECS}

_TABLE_STYLE = TableStyle(
    [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d1d5db")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f9fafb")]),
    ]
)


def _sheet_rows(spec: WidgetSpec, data: Any) -> list[list[str]]:
    rows = []
    for row in spec.rows(data):
        text_columns, rest = row[: spec.numeric_from], row[spec.numeric_from :]
        rows.append([escape_formula(v) for v in text_columns] + rest)
    return rows


def export_widget_excel(widget: str, data: Any) -> bytes:
    spec = ALL_WIDGET_SPECS[widget]
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = spec.title[:31]
    sheet.append(spec.headers)
    for row in _sheet_rows(spec, data):
        sheet.append(row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def export_widget_csv(widget: str, data: Any) -> bytes:
    spec = ALL_WIDGET_SPECS[widget]
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(spec.headers)
    for row in _sheet_rows(spec, data):
        writer.writerow(row)
    return output.getvalue().encode("utf-8")


def _pdf_header_story(meta: ExportMeta, styles) -> list:
    story: list = []
    if meta.logo_url:
        logo = logo_flowable(meta.logo_url, max_width=30 * mm, max_height=15 * mm)
        if logo:
            story.append(logo)
            story.append(Spacer(1, 2 * mm))
    story.append(Paragraph(xml_escape(meta.company_name), styles["Heading2"]))
    story.append(Paragraph(xml_escape(meta.report_title), styles["Title"]))
    story.append(Paragraph(f"Period: {xml_escape(meta.filter_label)}", styles["Normal"]))
    story.append(Paragraph(f"Generated: {meta.generated_at.strftime('%Y-%m-%d %H:%M')}", styles["Normal"]))
    story.append(Spacer(1, 4 * mm))
    return story


def _pdf_footer(canvas, doc) -> None:
    canvas.saveState()
    canvas.setFont("Helvetica", 8)
    canvas.setFillColor(colors.HexColor("#6b7280"))
    canvas.drawString(14 * mm, 10 * mm, f"Generated {datetime.now():%Y-%m-%d %H:%M}")
    canvas.drawRightString(landscape(A4)[0] - 14 * mm, 10 * mm, f"Page {doc.page}")
    canvas.restoreState()


def _widget_table_story(spec: WidgetSpec, data: Any, styles) -> list:
    story: list = [Paragraph(spec.title, styles["Heading2"]), Spacer(1, 2 * mm)]
    rows = spec.rows(data)
    if not rows:
        story.append(Paragraph("No data available for the selected period.", styles["Normal"]))
    else:
        table = Table([spec.headers] + rows, repeatRows=1)
        style = TableStyle(_TABLE_STYLE.getCommands())
        style.add("ALIGN", (spec.numeric_from, 0), (-1, -1), "RIGHT")
        table.setStyle(style)
        story.append(table)
    story.append(Spacer(1, 6 * mm))
    return story


def export_widget_pdf(widget: str, data: Any, meta: ExportMeta) -> bytes:
    spec = ALL_WIDGET_SPECS[widget]
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=14 * mm, bottomMargin=18 * mm, leftMargin=14 * mm, rightMargin=14 * mm)
    styles = getSampleStyleSheet()
    story = _pdf_header_story(meta, styles) + _widget_table_story(spec, data, styles)
    doc.build(story, onFirstPage=_pdf_footer, onLaterPages=_pdf_footer)
    return buffer.getvalue()


def export_full_analytics_excel(data: ProcurementAnalyticsOut) -> bytes:
    workbook = Workbook()
    first = True
    for spec in ANALYTICS_WIDGET_SPECS.values():
        sheet = workbook.active if first else workbook.create_sheet()
        first = False
        sheet.title = spec.title[:31]
        sheet.append(spec.headers)
        for row in _sheet_rows(spec, data):
            sheet.append(row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def export_full_analytics_pdf(data: ProcurementAnalyticsOut, meta: ExportMeta) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=14 * mm, bottomMargin=18 * mm, leftMargin=14 * mm, rightMargin=14 * mm)
    styles = getSampleStyleSheet()
    story = _pdf_header_story(meta, styles)
    for spec in ANALYTICS_WIDGET_SPECS.values():
        story.extend(_widget_table_story(spec, data, styles))
    doc.build(story, onFirstPage=_pdf_footer, onLaterPages=_pdf_footer)
    return buffer.getvalue()


def export_full_reports_excel(data: ProcurementReportsOut) -> bytes:
    workbook = Workbook()
    first = True
    for spec in REPORT_WIDGET_SPECS.values():
        sheet = workbook.active if first else workbook.create_sheet()
        first = False
        sheet.title = spec.title[:31]
        sheet.append(spec.headers)
        for row in _sheet_rows(spec, data):
            sheet.append(row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def export_full_reports_pdf(data: ProcurementReportsOut, meta: ExportMeta) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=14 * mm, bottomMargin=18 * mm, leftMargin=14 * mm, rightMargin=14 * mm)
    styles = getSampleStyleSheet()
    story = _pdf_header_story(meta, styles)
    for spec in REPORT_WIDGET_SPECS.values():
        story.extend(_widget_table_story(spec, data, styles))
    doc.build(story, onFirstPage=_pdf_footer, onLaterPages=_pdf_footer)
    return buffer.getvalue()
