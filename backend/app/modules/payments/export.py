from __future__ import annotations

import csv
import io

from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.core.sanitize import escape_formula
from app.schemas.payment import AgeingInvoiceOut, CustomerLedgerOut

LEDGER_HEADERS = ["Date", "Type", "Reference", "Description", "Debit", "Credit", "Balance"]
AGEING_HEADERS = ["Invoice #", "Customer", "Due Date", "Outstanding", "Days Overdue", "Bucket"]
OUTSTANDING_HEADERS = ["Invoice #", "Customer", "Due Date", "Outstanding", "Days Overdue"]

_TABLE_STYLE = TableStyle(
    [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d1d5db")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f9fafb")]),
    ]
)


def _pdf_report(title: str, headers: list[str], rows: list[list[str]], text_columns: int) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4), topMargin=14 * mm, bottomMargin=14 * mm, leftMargin=14 * mm, rightMargin=14 * mm
    )
    styles = getSampleStyleSheet()
    story = [Paragraph(title, styles["Title"]), Spacer(1, 4 * mm)]
    table = Table([headers] + rows, repeatRows=1)
    style = TableStyle(_TABLE_STYLE.getCommands() + [("ALIGN", (text_columns, 0), (-1, -1), "RIGHT")])
    table.setStyle(style)
    story.append(table)
    doc.build(story)
    return buffer.getvalue()


def _ledger_rows(ledger: CustomerLedgerOut) -> list[list[str]]:
    return [
        [
            entry.date.strftime("%Y-%m-%d %H:%M"),
            entry.type.title(),
            entry.reference,
            entry.description,
            f"{entry.debit:.2f}" if entry.debit else "",
            f"{entry.credit:.2f}" if entry.credit else "",
            f"{entry.balance:.2f}",
        ]
        for entry in ledger.entries
    ]


def export_ledger_excel(ledger: CustomerLedgerOut) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Ledger"
    sheet.append(LEDGER_HEADERS)
    for row in _ledger_rows(ledger):
        text_columns, rest = row[:4], row[4:]
        sheet.append([escape_formula(value) for value in text_columns] + rest)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def export_ledger_csv(ledger: CustomerLedgerOut) -> bytes:
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(LEDGER_HEADERS)
    for row in _ledger_rows(ledger):
        text_columns, rest = row[:4], row[4:]
        writer.writerow([escape_formula(value) for value in text_columns] + rest)
    return output.getvalue().encode("utf-8")


def export_ledger_pdf(ledger: CustomerLedgerOut) -> bytes:
    return _pdf_report(f"Customer Ledger — {ledger.customer_name}", LEDGER_HEADERS, _ledger_rows(ledger), text_columns=4)


def _ageing_rows(invoices: list[AgeingInvoiceOut]) -> list[list[str]]:
    return [
        [
            inv.invoice_number,
            inv.customer_name or "Walk-in",
            inv.due_date.strftime("%Y-%m-%d") if inv.due_date else "",
            f"{inv.outstanding_amount:.2f}",
            str(inv.days_overdue),
            inv.bucket,
        ]
        for inv in invoices
    ]


def export_ageing_excel(report) -> bytes:  # noqa: ANN001 - AgeingReportOut, kept loose to avoid an import cycle in type hints
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Ageing"
    sheet.append(AGEING_HEADERS)
    for row in _ageing_rows(report.invoices):
        text_columns, rest = row[:3], row[3:]
        sheet.append([escape_formula(value) for value in text_columns] + rest)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def export_ageing_csv(report) -> bytes:  # noqa: ANN001
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(AGEING_HEADERS)
    for row in _ageing_rows(report.invoices):
        text_columns, rest = row[:3], row[3:]
        writer.writerow([escape_formula(value) for value in text_columns] + rest)
    return output.getvalue().encode("utf-8")


def export_ageing_pdf(report) -> bytes:  # noqa: ANN001
    return _pdf_report("Ageing Report", AGEING_HEADERS, _ageing_rows(report.invoices), text_columns=3)


def _outstanding_rows(invoices: list[AgeingInvoiceOut]) -> list[list[str]]:
    return [
        [
            inv.invoice_number,
            inv.customer_name or "Walk-in",
            inv.due_date.strftime("%Y-%m-%d") if inv.due_date else "",
            f"{inv.outstanding_amount:.2f}",
            str(inv.days_overdue),
        ]
        for inv in invoices
    ]


def export_outstanding_excel(invoices: list[AgeingInvoiceOut]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Outstanding"
    sheet.append(OUTSTANDING_HEADERS)
    for row in _outstanding_rows(invoices):
        text_columns, rest = row[:3], row[3:]
        sheet.append([escape_formula(value) for value in text_columns] + rest)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def export_outstanding_csv(invoices: list[AgeingInvoiceOut]) -> bytes:
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(OUTSTANDING_HEADERS)
    for row in _outstanding_rows(invoices):
        text_columns, rest = row[:3], row[3:]
        writer.writerow([escape_formula(value) for value in text_columns] + rest)
    return output.getvalue().encode("utf-8")


def export_outstanding_pdf(invoices: list[AgeingInvoiceOut]) -> bytes:
    return _pdf_report("Outstanding Report", OUTSTANDING_HEADERS, _outstanding_rows(invoices), text_columns=3)
