from __future__ import annotations

import io

from openpyxl import load_workbook

from app.modules.inventory.parsers.base import ImportedRow, ParserError


class ExcelParser:
    async def parse(self, file_bytes: bytes, filename: str) -> list[ImportedRow]:
        try:
            workbook = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
            sheet = workbook.active
            rows_iter = sheet.iter_rows(values_only=True)
            headers = next(rows_iter, None)
        except Exception as exc:  # noqa: BLE001 - any openpyxl/zip failure means an unreadable file
            raise ParserError(f"Could not read Excel file: {exc}") from exc

        if not headers or all(h is None for h in headers):
            raise ParserError("The uploaded file has no header row")

        header_names = [str(h).strip() if h is not None else "" for h in headers]

        rows: list[ImportedRow] = []
        for values in rows_iter:
            if values is None or all(v is None or str(v).strip() == "" for v in values):
                continue
            raw = {
                header_names[i]: ("" if value is None else str(value).strip())
                for i, value in enumerate(values)
                if i < len(header_names) and header_names[i]
            }
            rows.append(ImportedRow(raw=raw))
        return rows
